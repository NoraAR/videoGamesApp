from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
import pandas as pd
from io import BytesIO

def download_excel(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

def create_excel(year):
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Añadir texto a una celda
    # Añadir texto a una celda
    ws['E1'] = f"Reporte Personalizado {year}"

    # Establecer fuente, tamaño y color
    ws['E1'].font = Font(name='Amercian Typewriter', size=20, bold=True, italic=True, color='FFFFFF')

    # Ajustar el tamaño de la celda
    ws.row_dimensions[1].height = 60
    ws.column_dimensions['E'].width = 60

    # Alinear el texto
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar un color de fondo a una celda
    ws['E1'].fill = PatternFill(start_color='050505', end_color='050505', fill_type='solid')

    # Agregar la primera imagen
    img1 = Image(f'Images/pie_chart_genres_{year}.png')
    ws.add_image(img1, 'A6')  # Cambiar la posición vertical de la primera imagen

    # Agregar la segunda imagen
    img2 = Image(f'Images/bar_chart_top_5_{year}.png')
    ws.add_image(img2, 'F6')  # Cambiar la posición vertical de la segunda imagen

    # Guardar el libro de trabajo
    wb.save('Reports/reporte_personalizado.xlsx')

    return wb
